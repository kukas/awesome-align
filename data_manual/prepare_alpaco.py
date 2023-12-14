import openpyxl
from xml.etree.ElementTree import Element, SubElement, tostring, ElementTree
from xml.dom import minidom
import sys
import os

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from data.tokenize import get_tokenizer

def create_xml_element(tag, text):
    element = Element(tag)
    element.text = text
    return element

def convert_xlsx_to_xml(input_file, output_file, base_id):
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active

    sentences = Element("sentences")

    cs_tokenizer = get_tokenizer("cs")
    uk_tokenizer = get_tokenizer("uk")

    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_col=sheet.max_column, values_only=True), start=1):
        lang1_text, lang2_text = row
        if lang1_text is None or lang2_text is None:
            continue

        lang1_tokens = cs_tokenizer(lang1_text)
        lang2_tokens = uk_tokenizer(lang2_text)
        lang1_text_tokenized = " ".join(lang1_tokens)
        lang2_text_tokenized = " ".join(lang2_tokens)

        # check if the only difference between the original and tokenization is spaces
        assert lang1_text_tokenized.replace(" ", "") == lang1_text.replace(" ", "")
        assert lang2_text_tokenized.replace(" ", "") == lang2_text.replace(" ", "")

        s_element = Element("s", id=f"{base_id}-s{row_number}")

        lang1_element = create_xml_element("czech", lang1_text_tokenized)
        lang2_element = create_xml_element("ukrainian", lang2_text_tokenized)
        sure_element = create_xml_element("sure", "")
        possible_element = create_xml_element("possible", "")
        phrasal_element = create_xml_element("phrasal", "")

        s_element.append(lang1_element)
        s_element.append(lang2_element)
        s_element.append(sure_element)
        s_element.append(possible_element)
        s_element.append(phrasal_element)

        sentences.append(s_element)

    tree = ElementTree(sentences)
    tree_str = tostring(sentences, encoding="utf-8").decode("utf-8")
    pretty_xml = minidom.parseString(tree_str).toprettyxml(indent="  ")

    with open(output_file, "w", encoding="utf-8") as xml_file:
        # unescape the xml...
        # the xml library escapes even the text inside the xml tags which
        # is something we do not want
        pretty_xml = pretty_xml.replace("&quot;", "\"")
        # of course this can lead to invalid xml, caution!
        xml_file.write(pretty_xml)

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python script.py input_excel_file output_xml_file")
        sys.exit(1)

    input_excel_file = sys.argv[1]
    output_xml_file = sys.argv[2]

    base_id = os.path.splitext(os.path.basename(output_xml_file))[0]

    convert_xlsx_to_xml(input_excel_file, output_xml_file, base_id)
